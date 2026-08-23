"""Reports for a multi-model panel run.

The panel report answers a different question from a single-model report. Not
"what is the verdict" but "what did several independent analysts conclude, where
did they split, what changed when they read each other, and how much is the
agreement actually worth". So it leads with the disagreements.
"""
from __future__ import annotations

import html
import json
from pathlib import Path
from typing import Any, List

from ..console import out as _out
from .common import ASSESSMENT_WORD, as_list, score_color, theme as get_theme, txt


def _e(v: Any) -> str:
    return html.escape(str(v if v not in (None, "") else "—"))


# =================================================================== markdown

def build_panel_markdown(result, lens: str) -> str:
    cons = result.consensus.get(lens, {})
    metrics = result.metrics.get(lens, {})
    working = result.working
    L: List[str] = []
    add = L.append

    verdict = cons.get("consensus_verdict") or {}
    add(f"# {result.company} — Panel Analysis")
    add("")
    add(f"**{len(working)} AI analysts · {lens} lens · "
        f"{result.stats.get('rounds', 0)} cross-review round(s)**")
    add("")
    if cons.get("headline"):
        add(f"> {cons['headline']}")
        add("")
    add("| | |")
    add("|---|---|")
    add(f"| **Consensus verdict** | {txt(verdict.get('call'))} |")
    add(f"| **Agreement** | {txt(verdict.get('agreement'))} |")
    add(f"| **Confidence** | {txt(verdict.get('confidence'))} |")
    add(f"| **Score spread** | {txt((metrics.get('score') or {}).get('spread'))} points "
        f"({txt((metrics.get('score') or {}).get('convergence'))}) |")
    add(f"| **Positions changed after review** | "
        f"{txt(metrics.get('total_position_changes'))} |")
    add(f"| **Panel** | {', '.join(p.name for p in working)} |")
    add(f"| **Review rounds** | {txt(result.stats.get('rounds_run'))} run "
        f"({txt(result.stats.get('strategy'))} stopping rule) |")
    add(f"| **Chair** | {txt(result.stats.get('chair'))} |")
    add(f"| **Generated** | {txt(result.stats.get('generated_at'))} |")
    add("")
    if verdict.get("rationale"):
        add(f"*{verdict['rationale']}*")
        add("")

    # ------------------------------------------------------- the split view
    add("## Where the panel landed")
    add("")
    add("| Panelist | Model | Verdict | Score | Changed after review |")
    add("|---|---|---|:--:|---|")
    for m in metrics.get("movement") or []:
        moved = ("verdict changed: "
                 f"{txt(m.get('verdict_before'))} → {txt(m.get('verdict_after'))}"
                 if m.get("verdict_before") != m.get("verdict_after")
                 else f"{txt(m.get('changes'))} position(s), verdict held")
        drift = ""
        if m.get("score_before") != m.get("score_after"):
            drift = f" ({txt(m.get('score_before'))} → {txt(m.get('score_after'))})"
        add(f"| {txt(m.get('panelist'))} | {txt(m.get('name'))} | "
            f"{txt(m.get('verdict_after'))} | {txt(m.get('score_after'))}{drift} | {moved} |")
    add("")

    failed = result.stats.get("panelists_failed") or []
    if failed:
        add("**Panelists that failed to run:**")
        add("")
        for f in failed:
            add(f"- `{txt(f.get('name'))}` — {txt(f.get('error'))}")
        add("")

    if result.stats.get("stopped_because"):
        add(f"*Why the panel stopped when it did: {result.stats['stopped_because']}*")
        add("")

    add(cons.get("summary") or "")
    add("")

    # ---------------------------------------------------- the ranked reports
    vote = result.votes.get(lens)
    if vote and vote.ballots:
        add("## The individual reports, ranked")
        add("")
        add("The synthesis above is a committee document, and a committee document can "
            "smooth away the disagreement that is the most useful thing here. So each "
            "panelist's own report is kept intact and listed below, ordered by how the "
            "rest of the panel ranked it.")
        add("")
        add("Panelists ranked each other on whether conclusions follow from the evidence "
            "cited, not on whether they agreed. Nobody ranked themselves.")
        add("")
        add("| Rank | Report | Score | First-place votes | Verdict |")
        add("|:--:|---|:--:|:--:|---|")
        for i, label in enumerate(vote.order, 1):
            p_ = next((x for x in working if x.label == label), None)
            verdict = ((p_.final(lens).get("verdict") or {}).get("call", "—")
                       if p_ else "—")
            mark = " ⭐" if label == vote.winner else ""
            add(f"| {i}{mark} | **{label}** — {txt(p_.name if p_ else '')} | "
                f"{txt(vote.scores.get(label))} | {txt(vote.firsts.get(label, 0))} | "
                f"{verdict} |")
        add("")
        add(f"*{vote.note}*")
        add("")
        if not vote.decisive:
            add("Because the vote was tied, no single report is presented as the "
                "panel's preferred analysis. Read the tied reports side by side.")
            add("")
        reasons = [(b.voter, t, r) for b in vote.ballots for t, r in b.reasons.items()]
        if reasons:
            add("<details><summary>Why each panelist ranked the others as it did</summary>")
            add("")
            for voter, target, reason in reasons[:24]:
                add(f"- **{voter}** on {target}: {reason}")
            add("")
            add("</details>")
            add("")
        notes = [b.note for b in vote.ballots if b.note]
        if notes:
            add("**What the panel thought it collectively missed:**")
            add("")
            for n in notes:
                add(f"- {n}")
            add("")

    # --------------------------------------------------------- agreement
    agree = cons.get("where_all_agree") or []
    if agree:
        add("## What every panelist agreed on")
        add("")
        add("These points survived independent analysis by every model on the panel, "
            "and survived cross-examination by the others.")
        add("")
        for a in agree:
            if isinstance(a, dict):
                add(f"- **{txt(a.get('point'))}** — {txt(a.get('why_it_is_robust'))}")
            else:
                add(f"- {a}")
        add("")

    # --------------------------------------------------------- contested
    contested = cons.get("contested") or []
    if contested:
        add("## Where the panel split")
        add("")
        add("Disagreement here is the most useful output of a panel: it marks exactly "
            "where the evidence is thin enough that competent analysts diverge.")
        add("")
        for c in contested:
            add(f"### {txt(c.get('topic'))}")
            add("")
            for pos in c.get("positions") or []:
                add(f"- **{txt(pos.get('panelist'))}**: {txt(pos.get('position'))}  ")
                add(f"  *evidence: {txt(pos.get('evidence_quality'))}"
                    + (f" — {', '.join(str(s) for s in as_list(pos.get('source_ids')))}"
                       if pos.get("source_ids") else "") + "*")
            add("")
            if c.get("resolution"):
                add(f"**Where the evidence points:** {c['resolution']}")
                add("")
            if c.get("what_would_settle_it"):
                add(f"**What would settle it:** {c['what_would_settle_it']}")
                add("")

    # ------------------------------------------------- claim-level agreement
    claims = metrics.get("claims") or []
    if claims:
        add("## Claim-by-claim, across the panel")
        add("")
        labels = [p.label for p in working]
        add("Claims are matched across panelists by content — the numbers they quote "
            "and the words they use — not by each panelist's own numbering, which is "
            "independent and not comparable.")
        add("")
        add("| Claim | " + " | ".join(labels) + " | Consensus |")
        add("|---|" + "|".join([":--:"] * len(labels)) + "|---|")
        for c in claims:
            row = [f"**{txt(c.get('id'))}** {txt(c.get('claim'))[:70]}"]
            for lbl in labels:
                a = (c.get("assessments") or {}).get(lbl)
                row.append(ASSESSMENT_WORD.get(a, a) if a else "*not raised*")
            if c.get("single_panelist"):
                verdict = "only one panelist raised this"
            elif c.get("unanimous"):
                verdict = "unanimous"
            elif c.get("contested"):
                verdict = "no consensus ⚠"
            else:
                verdict = f"agreed by {c.get('raised_by')} of {c.get('of_panelists')}"
            row.append(verdict)
            add("| " + " | ".join(row) + " |")
        add("")
        contested_ids = metrics.get("contested_claims") or []
        if contested_ids:
            add(f"⚠ marks the {len(contested_ids)} claim(s) the panel assessed "
                f"differently: {', '.join(contested_ids)}. Those are the ones to check "
                f"yourself.")
            add("")
        solo = metrics.get("single_panelist_claims") or []
        if solo:
            add(f"*{len(solo)} claim(s) were raised by only one panelist. That is not "
                f"disagreement — the others did not address them at all, which usually "
                f"means the claim was easy to miss.*")
            add("")

    # ------------------------------------------------------- dimension spread
    dims = metrics.get("dimensions") or {}
    if dims:
        add("## Scorecard, across the panel")
        add("")
        add("| Dimension | Mean | Spread | Individual scores |")
        add("|---|:--:|:--:|---|")
        for d, st in sorted(dims.items(), key=lambda kv: -kv[1]["spread"]):
            flag = " ⚠" if st.get("contested") else ""
            add(f"| {d}{flag} | {st['mean']} | {st['spread']} | "
                f"{', '.join(str(s) for s in st['scores'])} |")
        add("")
        add("*A wide spread means the panelists read the same evidence and weighted it "
            "differently — usually a sign that the deck left the question open.*")
        add("")

    # -------------------------------------------------------- minority view
    minority = cons.get("minority_report") or []
    if minority:
        add("## Minority report")
        add("")
        add("Dissenting positions, stated at their strongest. A dissent that turns out "
            "to be right is the most valuable thing a panel can produce.")
        add("")
        for m in minority:
            add(f"### {txt(m.get('panelist'))}")
            add("")
            add(txt(m.get("position")))
            add("")
            if m.get("why_it_deserves_a_hearing"):
                add(f"*{m['why_it_deserves_a_hearing']}*")
                add("")

    # ------------------------------------------------------ what changed
    add("## What changed when the panelists read each other")
    add("")
    any_change = False
    for p in working:
        changes = (p.review or {}).get("position_changes") or []
        held = (p.review or {}).get("positions_held") or []
        if not changes and not held:
            continue
        any_change = True
        add(f"### {p.label} — {p.name}")
        add("")
        if changes:
            add("**Conceded:**")
            add("")
            for ch in changes:
                add(f"- {txt(ch.get('what_changes'))}  ")
                add(f"  *{txt(ch.get('from'))}* → *{txt(ch.get('to'))}*  ")
                add(f"  prompted by {txt(ch.get('prompted_by'))}: {txt(ch.get('evidence'))}")
            add("")
        if held:
            add("**Held despite challenge:**")
            add("")
            for hd in held:
                add(f"- {txt(hd.get('position'))} — {txt(hd.get('why_you_hold_it'))}")
            add("")
        if (p.review or {}).get("self_assessment"):
            add(f"**Self-assessment:** {p.review['self_assessment']}")
            add("")
    if not any_change:
        add("No panelist changed position after reading the others. That can mean the "
            "analyses were genuinely robust — or that they shared the same blind spots. "
            "Read the reliability note below before treating it as confirmation.")
        add("")

    # ------------------------------------------------------- how it stopped
    if result.round_log:
        add("<details><summary>How the panel decided to stop</summary>")
        add("")
        add("| After round | Spread | Agreement | Changes | Contested | Continue? | Why |")
        add("|:--:|:--:|---|:--:|:--:|:--:|---|")
        for e in result.round_log:
            add(f"| {txt(e.get('after_round'))} | {txt(e.get('spread'))} | "
                f"{txt(e.get('verdict_agreement'))} | {txt(e.get('position_changes'))} | "
                f"{txt(e.get('contested_claims'))} | "
                f"{'yes' if e.get('proceed') else 'stop'} | {txt(e.get('reason'))} |")
        add("")
        add("</details>")
        add("")

    # -------------------------------------------------------- reliability
    rel = cons.get("reliability") or {}
    if rel:
        add("## How much this agreement is worth")
        add("")
        if rel.get("what_agreement_means_here"):
            add(rel["what_agreement_means_here"])
            add("")
        blind = as_list(rel.get("shared_blind_spots"))
        if blind:
            add("**Blind spots the whole panel may share:**")
            add("")
            for b in blind:
                add(f"- {b}")
            add("")
        if rel.get("caution"):
            add(f"> {rel['caution']}")
            add("")
    add("Models trained on overlapping data, reading the same sources, will agree for "
        "correlated reasons. Panel agreement raises confidence; it does not establish "
        "fact. Every figure below is traceable to a numbered source — check the ones "
        "that matter to your decision.")
    add("")

    # -------------------------------------------- individual final reports
    add("---")
    add("")
    add("## Annex — each panelist's final analysis")
    add("")
    primary = result.primary_result()
    ordered = (sorted(working, key=lambda x: (x.rank or 99))
               if any(x.rank for x in working) else working)
    for p in ordered:
        rank = f" — ranked #{p.rank} by the panel" if p.rank else ""
        add(f"### {p.label} — {p.name}{rank}")
        add("")
        final = p.final(lens)
        v = final.get("verdict") or {}
        add(f"**{txt(v.get('call'))}** · {txt(v.get('confidence'))} confidence")
        add("")
        if final.get("headline"):
            add(f"> {final['headline']}")
            add("")
        add(txt(final.get("summary")))
        add("")
        log = ((final.get("_meta") or {}).get("revision_log")) or []
        if log:
            add("<details><summary>What this panelist revised</summary>")
            add("")
            for entry in log:
                add(f"- **{txt(entry.get('field'))}**: {txt(entry.get('from'))} → "
                    f"{txt(entry.get('to'))} — {txt(entry.get('reason'))} "
                    f"({txt(entry.get('prompted_by'))})")
            add("")
            add("</details>")
            add("")

    # -------------------------------------------------- shared annexes
    if primary is not None:
        from .markdown_renderer import _references_markdown, _security_markdown

        add("---")
        add("")
        add(_references_markdown(primary))
        add(_security_markdown(primary))

    add("---")
    add("")
    add(f"*Generated by DeckScope · panel of "
        f"{len(working)} · chaired by {result.stats.get('chair', '?')}. AI-generated "
        f"analysis: verify every figure against its cited source before relying on it. "
        f"Not investment advice.*")
    return "\n".join(L)


# ======================================================================= html

def build_panel_html(result, lens: str, theme_name: str = "slate") -> str:
    t = get_theme(theme_name)
    cons = result.consensus.get(lens, {})
    metrics = result.metrics.get(lens, {})
    working = result.working
    verdict = cons.get("consensus_verdict") or {}
    P: List[str] = []
    add = P.append

    add(f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{_e(result.company)} — Panel Analysis</title><style>
:root{{--accent:{t['accent']};--ink:{t['ink']};--muted:{t['muted']};--bg:{t['bg']};
--panel:{t['panel']};--line:{t['line']};--good:{t['good']};--warn:{t['warn']};--bad:{t['bad']}}}
*{{box-sizing:border-box}}
body{{margin:0;background:var(--bg);color:var(--ink);
font:16px/1.65 -apple-system,BlinkMacSystemFont,"Segoe UI",Inter,Roboto,sans-serif}}
.wrap{{max-width:1000px;margin:0 auto;padding:48px 24px 96px}}
header{{border-bottom:3px solid var(--accent);padding-bottom:24px;margin-bottom:30px}}
.eyebrow{{text-transform:uppercase;letter-spacing:.12em;font-size:12px;color:var(--muted);font-weight:600}}
h1{{font-size:34px;line-height:1.2;margin:8px 0 4px;letter-spacing:-.02em}}
h2{{font-size:22px;margin:46px 0 14px;padding-bottom:8px;border-bottom:1px solid var(--line)}}
h3{{font-size:17px;margin:26px 0 8px}}
.headline{{font-size:19px;line-height:1.5;background:var(--panel);
border-left:4px solid var(--accent);padding:16px 20px;border-radius:0 6px 6px 0;margin:20px 0}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;margin:20px 0}}
.stat{{background:var(--panel);border:1px solid var(--line);border-radius:8px;padding:14px 16px}}
.stat .k{{font-size:11px;text-transform:uppercase;letter-spacing:.08em;color:var(--muted);font-weight:600}}
.stat .v{{font-size:19px;font-weight:650;margin-top:4px;line-height:1.25}}
table{{width:100%;border-collapse:collapse;margin:16px 0;font-size:14px}}
th,td{{text-align:left;padding:9px 11px;border-bottom:1px solid var(--line);vertical-align:top}}
th{{font-size:11.5px;text-transform:uppercase;letter-spacing:.07em;color:var(--muted);font-weight:650}}
.tw{{overflow-x:auto}}
.panel-card{{border:1px solid var(--line);border-radius:9px;padding:16px 20px;
margin:12px 0;background:var(--panel)}}
.split{{border-left:4px solid var(--warn)}}
.agree{{border-left:4px solid var(--good)}}
.dissent{{border-left:4px solid var(--bad)}}
.tag{{display:inline-block;font-size:11px;font-weight:700;letter-spacing:.05em;
text-transform:uppercase;padding:3px 9px;border-radius:20px;color:#fff}}
.t-unanimous{{background:var(--good)}}.t-majority{{background:var(--warn)}}
.t-split{{background:var(--bad)}}.t-irreconcilable{{background:var(--bad)}}
.contested{{color:var(--bad);font-weight:650}}
.bar{{height:7px;background:var(--line);border-radius:4px;overflow:hidden;min-width:70px;margin-top:5px}}
.bar span{{display:block;height:100%;border-radius:4px}}
ul{{padding-left:20px}}li{{margin:5px 0}}
details{{margin:12px 0}}summary{{cursor:pointer;color:var(--accent);font-weight:600}}
a{{color:var(--accent)}}
footer{{margin-top:60px;padding-top:20px;border-top:1px solid var(--line);
font-size:12.5px;color:var(--muted)}}
@media print{{.wrap{{max-width:none;padding:0}}h2{{page-break-after:avoid}}}}
</style></head><body><div class="wrap">""")

    add(f"""<header><div class="eyebrow">Panel of {len(working)} · {_e(lens)} lens ·
{_e(result.stats.get('rounds', 0))} cross-review round(s)</div>
<h1>{_e(result.company)}</h1>
<div style="color:var(--muted)">Several AI analysts, working independently, then
reviewing each other</div></header>""")

    if cons.get("headline"):
        add(f'<div class="headline">{_e(cons["headline"])}</div>')

    agreement = str(verdict.get("agreement") or "split").lower().split()[0]
    add('<div class="grid">')
    add(f'<div class="stat"><div class="k">Consensus verdict</div>'
        f'<div class="v">{_e(verdict.get("call"))}</div></div>')
    add(f'<div class="stat"><div class="k">Agreement</div><div class="v">'
        f'<span class="tag t-{html.escape(agreement)}">{_e(verdict.get("agreement"))}</span>'
        f'</div></div>')
    add(f'<div class="stat"><div class="k">Score spread</div><div class="v">'
        f'{_e((metrics.get("score") or {}).get("spread"))} pts '
        f'<span style="font-size:13px;color:var(--muted)">'
        f'({_e((metrics.get("score") or {}).get("convergence"))})</span></div></div>')
    add(f'<div class="stat"><div class="k">Positions changed</div>'
        f'<div class="v">{_e(metrics.get("total_position_changes"))}</div></div>')
    add("</div>")

    if verdict.get("rationale"):
        add(f'<p style="color:var(--muted)">{_e(verdict["rationale"])}</p>')

    # panel table
    add("<h2>Where the panel landed</h2><div class='tw'><table>"
        "<tr><th>Panelist</th><th>Model</th><th>Verdict</th><th>Score</th>"
        "<th>After cross-review</th></tr>")
    for m in metrics.get("movement") or []:
        changed = m.get("verdict_before") != m.get("verdict_after")
        note = (f'<span class="contested">verdict changed</span>: '
                f'{_e(m.get("verdict_before"))} → {_e(m.get("verdict_after"))}'
                if changed else
                f'{_e(m.get("changes"))} position(s) conceded, verdict held')
        drift = ""
        if m.get("score_before") != m.get("score_after"):
            drift = (f'<div style="font-size:12px;color:var(--muted)">'
                     f'{_e(m.get("score_before"))} → {_e(m.get("score_after"))}</div>')
        add(f"<tr><td><b>{_e(m.get('panelist'))}</b></td><td>{_e(m.get('name'))}</td>"
            f"<td>{_e(m.get('verdict_after'))}</td>"
            f"<td>{_e(m.get('score_after'))}{drift}</td><td>{note}</td></tr>")
    add("</table></div>")

    failed = result.stats.get("panelists_failed") or []
    if failed:
        add("<div class='panel-card dissent'><b>Panelists that failed to run</b><ul>"
            + "".join(f"<li><code>{_e(f.get('name'))}</code> — {_e(f.get('error'))}</li>"
                      for f in failed) + "</ul></div>")

    add("<h2>Summary</h2>")
    for para in (cons.get("summary") or "").split("\n"):
        if para.strip():
            add(f"<p>{_e(para.strip())}</p>")

    agree_pts = cons.get("where_all_agree") or []
    if agree_pts:
        add("<h2>What every panelist agreed on</h2>")
        add("<p style='color:var(--muted)'>These points survived independent analysis by "
            "every model, and survived cross-examination by the others.</p>")
        for a in agree_pts:
            if isinstance(a, dict):
                add(f'<div class="panel-card agree"><b>{_e(a.get("point"))}</b><br>'
                    f'<span style="color:var(--muted);font-size:14px">'
                    f'{_e(a.get("why_it_is_robust"))}</span></div>')
            else:
                add(f'<div class="panel-card agree">{_e(a)}</div>')

    contested = cons.get("contested") or []
    if contested:
        add("<h2>Where the panel split</h2>")
        add("<p style='color:var(--muted)'>Disagreement is the most useful output of a "
            "panel: it marks exactly where the evidence is thin enough that competent "
            "analysts diverge.</p>")
        for c in contested:
            add(f'<div class="panel-card split"><h3 style="margin-top:0">'
                f'{_e(c.get("topic"))}</h3>')
            for pos in c.get("positions") or []:
                add(f'<p><b>{_e(pos.get("panelist"))}:</b> {_e(pos.get("position"))}<br>'
                    f'<span style="color:var(--muted);font-size:13px">evidence: '
                    f'{_e(pos.get("evidence_quality"))}'
                    + (f' · {_e(", ".join(str(s) for s in as_list(pos.get("source_ids"))))}'
                       if pos.get("source_ids") else "")
                    + '</span></p>')
            if c.get("resolution"):
                add(f'<p><b>Where the evidence points:</b> {_e(c["resolution"])}</p>')
            if c.get("what_would_settle_it"):
                add(f'<p><b>What would settle it:</b> {_e(c["what_would_settle_it"])}</p>')
            add("</div>")

    claims = metrics.get("claims") or []
    if claims:
        labels = [p.label for p in working]
        add("<h2>Claim-by-claim, across the panel</h2>")
        add("<p style='color:var(--muted)'>Claims are matched across panelists by "
            "content — the numbers they quote and the words they use — not by each "
            "panelist's own numbering, which is independent and not comparable.</p>")
        add("<div class='tw'><table><tr><th>Claim</th>"
            + "".join(f"<th>{_e(l)}</th>" for l in labels) + "<th>Consensus</th></tr>")
        for c in claims:
            add(f"<tr><td><b>{_e(c.get('id'))}</b> {_e(str(c.get('claim'))[:80])}</td>")
            for lbl in labels:
                a = (c.get("assessments") or {}).get(lbl)
                cell = (_e(ASSESSMENT_WORD.get(a, a)) if a
                        else "<i style='color:var(--muted)'>not raised</i>")
                add(f"<td>{cell}</td>")
            if c.get("single_panelist"):
                label, cls = "only one panelist raised this", ""
            elif c.get("unanimous"):
                label, cls = "unanimous", ""
            elif c.get("contested"):
                label, cls = "no consensus", " class='contested'"
            else:
                label, cls = f"agreed by {c.get('raised_by')} of {c.get('of_panelists')}", ""
            add(f"<td{cls}>{_e(label)}</td></tr>")
        add("</table></div>")

    dims = metrics.get("dimensions") or {}
    if dims:
        add("<h2>Scorecard, across the panel</h2><div class='tw'><table>"
            "<tr><th>Dimension</th><th>Mean</th><th>Spread</th><th>Individual scores</th></tr>")
        for d, st in sorted(dims.items(), key=lambda kv: -kv[1]["spread"]):
            cls = " class='contested'" if st.get("contested") else ""
            bars = "".join(
                f'<div class="bar" title="{s}"><span style="width:{min(100, s*10)}%;'
                f'background:{score_color(s, t)}"></span></div>' for s in st["scores"])
            add(f"<tr><td{cls}>{_e(d)}</td><td>{_e(st['mean'])}</td>"
                f"<td{cls}>{_e(st['spread'])}</td>"
                f"<td>{_e(', '.join(str(s) for s in st['scores']))}{bars}</td></tr>")
        add("</table></div>")
        add("<p style='color:var(--muted);font-size:14px'>A wide spread means the "
            "panelists read the same evidence and weighted it differently — usually a "
            "sign that the deck left the question open.</p>")

    minority = cons.get("minority_report") or []
    if minority:
        add("<h2>Minority report</h2>")
        add("<p style='color:var(--muted)'>Dissenting positions at their strongest. A "
            "dissent that turns out to be right is the most valuable thing a panel "
            "produces.</p>")
        for m in minority:
            add(f'<div class="panel-card dissent"><b>{_e(m.get("panelist"))}</b>'
                f'<p>{_e(m.get("position"))}</p>'
                + (f'<p style="color:var(--muted);font-size:14px">'
                   f'{_e(m.get("why_it_deserves_a_hearing"))}</p>'
                   if m.get("why_it_deserves_a_hearing") else "") + "</div>")

    add("<h2>What changed when the panelists read each other</h2>")
    any_change = False
    for p in working:
        changes = (p.review or {}).get("position_changes") or []
        held = (p.review or {}).get("positions_held") or []
        if not changes and not held:
            continue
        any_change = True
        add(f'<div class="panel-card"><h3 style="margin-top:0">{_e(p.label)} — '
            f'{_e(p.name)}</h3>')
        if changes:
            add("<b>Conceded</b><ul>")
            for ch in changes:
                add(f"<li>{_e(ch.get('what_changes'))}<br>"
                    f"<span style='color:var(--muted);font-size:13.5px'>"
                    f"<i>{_e(ch.get('from'))}</i> → <i>{_e(ch.get('to'))}</i> · "
                    f"prompted by {_e(ch.get('prompted_by'))}: "
                    f"{_e(ch.get('evidence'))}</span></li>")
            add("</ul>")
        if held:
            add("<b>Held despite challenge</b><ul>")
            for hd in held:
                add(f"<li>{_e(hd.get('position'))} — "
                    f"<span style='color:var(--muted)'>{_e(hd.get('why_you_hold_it'))}"
                    f"</span></li>")
            add("</ul>")
        if (p.review or {}).get("self_assessment"):
            add(f"<p><b>Self-assessment:</b> {_e(p.review['self_assessment'])}</p>")
        add("</div>")
    if not any_change:
        add("<div class='panel-card'>No panelist changed position after reading the "
            "others. That can mean the analyses were genuinely robust — or that they "
            "shared the same blind spots. Read the reliability note below before "
            "treating it as confirmation.</div>")

    rel = cons.get("reliability") or {}
    add("<h2>How much this agreement is worth</h2>")
    if rel.get("what_agreement_means_here"):
        add(f"<p>{_e(rel['what_agreement_means_here'])}</p>")
    blind = as_list(rel.get("shared_blind_spots"))
    if blind:
        add("<p><b>Blind spots the whole panel may share:</b></p><ul>"
            + "".join(f"<li>{_e(b)}</li>" for b in blind) + "</ul>")
    if rel.get("caution"):
        add(f'<div class="panel-card dissent">{_e(rel["caution"])}</div>')
    add("<p style='color:var(--muted)'>Models trained on overlapping data, reading the "
        "same sources, will agree for correlated reasons. Panel agreement raises "
        "confidence; it does not establish fact. Every figure is traceable to a numbered "
        "source — check the ones that matter to your decision.</p>")

    add("<h2>Annex — each panelist's final analysis</h2>")
    for p in working:
        final = p.final(lens)
        v = final.get("verdict") or {}
        add(f'<div class="panel-card"><h3 style="margin-top:0">{_e(p.label)} — '
            f'{_e(p.name)}</h3><p><b>{_e(v.get("call"))}</b> · '
            f'{_e(v.get("confidence"))} confidence</p>')
        if final.get("headline"):
            add(f"<p><i>{_e(final['headline'])}</i></p>")
        for para in (final.get("summary") or "").split("\n"):
            if para.strip():
                add(f"<p>{_e(para.strip())}</p>")
        log = ((final.get("_meta") or {}).get("revision_log")) or []
        if log:
            add("<details><summary>What this panelist revised</summary><ul>"
                + "".join(f"<li><b>{_e(e.get('field'))}</b>: {_e(e.get('from'))} → "
                          f"{_e(e.get('to'))} — {_e(e.get('reason'))} "
                          f"({_e(e.get('prompted_by'))})</li>" for e in log)
                + "</ul></details>")
        add("</div>")

    primary = result.primary_result()
    if primary is not None:
        from .html_renderer import _references_html, _security_html
        add(_references_html(primary))
        add(_security_html(primary))

    add(f"""<footer>Generated by DeckScope · panel of {len(working)} ·
chaired by {_e(result.stats.get('chair'))} · {_e(result.stats.get('generated_at'))}.<br>
AI-generated analysis. Verify every figure against its cited source before relying on it.
Not investment advice.</footer></div></body></html>""")
    return "\n".join(P)


# =================================================================== dispatch

def render_panel(result, out_dir: Path, base: str, formats: List[str],
                 theme: str = "slate") -> List[str]:
    """Write the panel report in each requested format, plus each panelist's own."""
    from .registry import render as render_single, resolve

    out_dir = Path(out_dir)
    written: List[str] = []
    fmts = [resolve(f) for f in formats]
    if "json" not in fmts:
        fmts.append("json")

    for lens in result.lenses:
        for fmt in fmts:
            try:
                if fmt == "md":
                    p = out_dir / f"{base}_panel_{lens}.md"
                    p.write_text(build_panel_markdown(result, lens), encoding="utf-8")
                elif fmt == "html":
                    p = out_dir / f"{base}_panel_{lens}.html"
                    p.write_text(build_panel_html(result, lens, theme), encoding="utf-8")
                elif fmt == "txt":
                    from .text_renderer import markdown_to_text
                    p = out_dir / f"{base}_panel_{lens}.txt"
                    p.write_text(markdown_to_text(build_panel_markdown(result, lens)),
                                 encoding="utf-8")
                elif fmt == "json":
                    if lens != result.lenses[0]:
                        continue
                    p = out_dir / f"{base}_panel_full.json"
                    p.write_text(json.dumps(result.to_dict(), indent=2, default=str),
                                 encoding="utf-8")
                elif fmt == "xlsx":
                    if lens != result.lenses[0]:
                        continue
                    p = Path(_panel_xlsx(result, out_dir, base))
                else:
                    # pdf / docx / pptx: render the consensus through the single-report
                    # renderers by presenting the consensus as a comparison.
                    continue
                written.append(str(p))
            except Exception as exc:  # noqa: BLE001
                _out(f"[panel] could not write {fmt}: {exc}")

    # Each panelist's own final report, in the same formats the user asked for.
    single_fmts = [f for f in fmts if f in ("pdf", "docx", "pptx", "md", "html")]
    for p in result.working:
        if not p.result:
            continue
        for lens in result.lenses:
            if p.revised.get(lens):
                p.result.comparisons[lens] = p.revised[lens]
        slug = _slug(p.name)
        for fmt in single_fmts:
            try:
                written.extend(render_single(fmt, p.result, out_dir,
                                             f"{base}_{slug}", theme=theme))
            except Exception:  # noqa: BLE001
                continue
    return written


def _panel_xlsx(result, out_dir: Path, base: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF", size=10)
    head_fill = PatternFill("solid", fgColor="2E5C8A")
    wrap = Alignment(vertical="top", wrap_text=True)

    def sheet(title, headers, rows, widths, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title[:31]
        ws.append(headers)
        for cell in ws[1]:
            cell.font, cell.fill, cell.alignment = head_font, head_fill, wrap
        for r in rows:
            ws.append([txt(v, "") for v in r])
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
        ws.freeze_panes = "A2"

    rows = []
    for lens, m in result.metrics.items():
        for mv in m.get("movement") or []:
            rows.append([lens, mv.get("panelist"), mv.get("name"),
                         mv.get("verdict_before"), mv.get("verdict_after"),
                         mv.get("score_before"), mv.get("score_after"),
                         mv.get("changes"), mv.get("held")])
    sheet("Panel", ["Lens", "Panelist", "Model", "Verdict before", "Verdict after",
                    "Score before", "Score after", "Conceded", "Held"],
          rows, [12, 12, 26, 24, 24, 13, 13, 11, 9], first=True)

    rows = []
    for lens, m in result.metrics.items():
        for c in m.get("claims") or []:
            rows.append([lens, c.get("id"), c.get("claim"), c.get("type"),
                         json.dumps(c.get("assessments") or {}),
                         json.dumps(c.get("local_ids") or {}),
                         f"{c.get('raised_by')}/{c.get('of_panelists')}",
                         "yes" if c.get("unanimous") else "no",
                         c.get("distinct_positions")])
    sheet("Claim agreement", ["Lens", "Key", "Claim", "Type", "Per-panelist assessment",
                              "Each panelist's own ID", "Raised by", "Unanimous",
                              "Distinct positions"],
          rows, [10, 7, 52, 14, 46, 34, 11, 11, 11])

    rows = []
    for lens, m in result.metrics.items():
        for d, st in (m.get("dimensions") or {}).items():
            rows.append([lens, d, st.get("mean"), st.get("spread"),
                         ", ".join(str(x) for x in st.get("scores", [])),
                         "yes" if st.get("contested") else "no"])
    sheet("Score spread", ["Lens", "Dimension", "Mean", "Spread", "Scores", "Contested"],
          rows, [12, 30, 9, 9, 26, 11])

    rows = []
    for p in result.working:
        for ch in (p.review or {}).get("position_changes") or []:
            rows.append([p.label, p.name, "conceded", ch.get("what_changes"),
                         ch.get("from"), ch.get("to"), ch.get("prompted_by"),
                         ch.get("evidence")])
        for hd in (p.review or {}).get("positions_held") or []:
            rows.append([p.label, p.name, "held", hd.get("position"), "", "",
                         hd.get("challenged_by"), hd.get("why_you_hold_it")])
    sheet("Cross-review", ["Panelist", "Model", "Outcome", "What", "From", "To",
                           "Prompted by", "Reason"],
          rows, [12, 26, 11, 52, 40, 40, 18, 52])

    reg = result.registry
    rows = [[s.sid, s.status, s.title, s.url, s.published, s.reliability,
             "; ".join(s.cited_by), s.note] for s in (reg.sources if reg else [])]
    sheet("References", ["ID", "Status", "Title", "URL", "Published", "Reliability",
                         "Supports", "Note"], rows, [8, 14, 48, 52, 14, 16, 40, 34])

    path = out_dir / f"{base}_panel.xlsx"
    wb.save(str(path))
    return str(path)


def _slug(name: str) -> str:
    import re
    s = re.sub(r"[^a-zA-Z0-9]+", "_", str(name)).strip("_").lower()
    return s or "panelist"
