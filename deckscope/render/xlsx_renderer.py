"""Spreadsheet output: one workbook, one sheet per view, filterable."""
from __future__ import annotations

from pathlib import Path
from typing import Any, List

from .common import as_list, txt


def render(result, out_dir: Path, base: str, theme: str = "slate", **kw: Any) -> List[str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("Excel output needs openpyxl: pip install openpyxl") from None

    wb = Workbook()
    accent = "2E5C8A"
    head_font = Font(bold=True, color="FFFFFF", size=10)
    head_fill = PatternFill("solid", fgColor=accent)
    wrap = Alignment(vertical="top", wrap_text=True)

    def sheet(title: str, headers: List[str], rows: List[List[Any]], widths: List[int],
              first: bool = False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title[:31]
        ws.append(headers)
        for c in ws[1]:
            c.font, c.fill, c.alignment = head_font, head_fill, wrap
        for r in rows:
            ws.append([txt(v, "") for v in r])
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = wrap
        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
        return ws

    # ---- Overview
    over = [["Company", result.company]]
    for k, v in (result.stats or {}).items():
        over.append([k.replace("_", " ").title(), txt(v, "")])
    for lens, comp in result.comparisons.items():
        over.append([f"{lens} verdict", (comp.get("verdict") or {}).get("call")])
        over.append([f"{lens} confidence", (comp.get("verdict") or {}).get("confidence")])
        over.append([f"{lens} score",
                     ((comp.get("_meta") or {}).get("weighted_score") or {}).get("score")])
        over.append([f"{lens} headline", comp.get("headline")])
    sheet("Overview", ["Field", "Value"], over, [34, 110], first=True)

    # ---- Scorecard
    rows = []
    for lens, comp in result.comparisons.items():
        for r in comp.get("scorecard") or []:
            rows.append([lens, r.get("dimension"), r.get("score"), r.get("weight"),
                         r.get("rationale"), ", ".join(as_list(r.get("source_ids")))])
    sheet("Scorecard", ["Lens", "Dimension", "Score", "Weight", "Rationale", "Sources"],
          rows, [12, 26, 8, 8, 90, 18])

    # ---- Claim audit
    rows = []
    for lens, comp in result.comparisons.items():
        for c in comp.get("claim_audit") or []:
            rows.append([lens, c.get("id"), c.get("claim"), c.get("assessment"),
                         c.get("evidence_quality"), c.get("market_evidence"),
                         c.get("delta"), c.get("so_what"),
                         ", ".join(as_list(c.get("source_ids")))])
    sheet("Claim audit", ["Lens", "ID", "Claim", "Assessment", "Evidence quality",
                          "Market evidence", "Gap", "So what", "Sources"],
          rows, [12, 7, 46, 18, 15, 70, 46, 46, 18])

    # ---- Risks
    rows = []
    for lens, comp in result.comparisons.items():
        for r in comp.get("risks") or []:
            rows.append([lens, r.get("risk"), r.get("severity"), r.get("likelihood"),
                         r.get("mitigation_or_test")])
    sheet("Risks", ["Lens", "Risk", "Severity", "Likelihood", "Test or mitigation"],
          rows, [12, 60, 12, 12, 60])

    # ---- Competitors
    land = result.market.get("competitive_landscape") or {}
    rows = []
    for group in ("incumbents", "challengers"):
        for c in land.get(group) or []:
            rows.append([group[:-1].title(), c.get("name"), c.get("position"),
                         c.get("funding_or_scale"), c.get("threat_level"),
                         c.get("url"), ", ".join(as_list(c.get("source_ids")))])
    sheet("Competitors", ["Type", "Company", "Position", "Scale", "Threat", "URL",
                          "Sources"], rows, [14, 26, 55, 22, 12, 44, 16])

    # ---- Market sizing
    rows = [[e.get("value"), e.get("year"), e.get("methodology"), e.get("source"),
             e.get("url"), ", ".join(as_list(e.get("source_ids")))]
            for e in (result.market.get("sizing") or {}).get("tam_estimates", []) or []]
    sheet("Market sizing", ["Estimate", "Year", "Methodology", "Source", "URL", "IDs"],
          rows, [18, 10, 40, 34, 44, 14])

    # ---- References: the full bibliography
    reg = getattr(result, "registry", None)
    rows = [[s.sid, s.status, s.title, s.url, s.domain, s.published, s.reliability,
             s.query, "; ".join(s.cited_by), s.note]
            for s in (reg.sources if reg else [])]
    sheet("References", ["ID", "Status", "Title", "URL", "Domain", "Published",
                         "Reliability", "Found via query", "Supports", "Note"],
          rows, [8, 14, 48, 52, 22, 14, 16, 40, 40, 34])

    # ---- Security findings
    sec = getattr(result, "security", None) or {}
    rows = []
    for key in ("deck", "web_sources"):
        for f in (sec.get(key) or {}).get("findings", []) or []:
            rows.append([key, f.get("severity"), f.get("code"), f.get("where"),
                         f.get("detail"), f.get("action"), f.get("excerpt")])
    sheet("Security", ["Input", "Severity", "Code", "Where", "Detail", "Action",
                       "Excerpt"], rows, [14, 12, 22, 24, 70, 12, 60])

    p = out_dir / f"{base}.xlsx"
    wb.save(str(p))
    return [str(p)]
