"""Analyze a folder of decks into one comparison spreadsheet.

    python examples/batch_analysis.py ./decks ./reports

Each deck gets its own full report; a summary workbook lets you sort the whole
pipeline by score, verdict, or how much the market contradicted the pitch.
"""
from __future__ import annotations

import sys
import traceback
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from deckscope import analyze
from deckscope.ingest.loader import SUPPORTED_EXTENSIONS
from deckscope.security.report import SecurityAbort

LENS = "investor"


def main(deck_dir: str, out_dir: str) -> int:
    decks = sorted(p for p in Path(deck_dir).iterdir()
                   if p.suffix.lower() in SUPPORTED_EXTENSIONS)
    if not decks:
        print(f"No readable decks in {deck_dir}.")
        print(f"Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}")
        return 1

    print(f"Analyzing {len(decks)} deck(s)…\n")
    rows = []

    for i, deck in enumerate(decks, 1):
        print(f"[{i}/{len(decks)}] {deck.name}")
        try:
            result = analyze(str(deck), lens=LENS, formats=["html", "json"],
                             out_dir=str(Path(out_dir) / deck.stem),
                             security="balanced", verbose=False)
        except SecurityAbort as exc:
            print(f"    BLOCKED by the security screen")
            rows.append({"deck": deck.name, "company": "—", "verdict": "BLOCKED",
                         "confidence": "—", "score": None, "security": "critical",
                         "sources_cited": 0, "sources_total": 0, "contested": "",
                         "headline": str(exc)[:200]})
            continue
        except Exception:  # noqa: BLE001
            print(f"    FAILED\n{traceback.format_exc(limit=1)}")
            rows.append({"deck": deck.name, "company": "—", "verdict": "ERROR",
                         "confidence": "—", "score": None, "security": "—",
                         "sources_cited": 0, "sources_total": 0, "contested": "",
                         "headline": "analysis failed — see the console output"})
            continue

        comp = result.primary
        stats = result.registry.stats() if result.registry else {}
        contradicted = [c["id"] for c in (comp.get("claim_audit") or [])
                        if c.get("assessment") == "contradicted"]
        score = ((comp.get("_meta") or {}).get("weighted_score") or {}).get("score")

        rows.append({
            "deck": deck.name,
            "company": result.company,
            "verdict": (comp.get("verdict") or {}).get("call"),
            "confidence": (comp.get("verdict") or {}).get("confidence"),
            "score": score,
            "security": (result.security or {}).get("overall_risk"),
            "sources_cited": stats.get("cited", 0),
            "sources_total": stats.get("total", 0),
            "contested": ", ".join(contradicted),
            "headline": comp.get("headline", ""),
        })
        print(f"    {rows[-1]['verdict']}  {score}/100  "
              f"security: {rows[-1]['security']}")

    write_summary(rows, Path(out_dir) / "summary.xlsx")
    print(f"\nDone. Summary: {Path(out_dir) / 'summary.xlsx'}")
    return 0


def write_summary(rows, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        import csv
        csv_path = path.with_suffix(".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(rows[0]))
            w.writeheader()
            w.writerows(rows)
        print(f"(openpyxl not installed — wrote {csv_path} instead)")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Pipeline"
    headers = ["Deck", "Company", "Verdict", "Confidence", "Score",
               "Security", "Cited", "Sources", "Contradicted claims", "Headline"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="2E5C8A")

    for r in rows:
        ws.append([r["deck"], r["company"], r["verdict"], r["confidence"], r["score"],
                   r["security"], r["sources_cited"], r["sources_total"],
                   r["contested"], r["headline"]])

    for i, w in enumerate([26, 24, 22, 12, 8, 11, 8, 9, 22, 90], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{len(rows) + 1}"
    wb.save(str(path))


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
