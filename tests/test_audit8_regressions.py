"""Sixth external audit (safety/privacy): every finding pinned.

The two that matter most: NDA enforcement was CLI-deep, not engine-deep —
the library path would send deck-derived content to a hosted model that
the CLI would have refused; and `improve --nda` validated one provider
while the pipeline sends the whole deck to a second one. Both now fail
closed at the layer that actually makes the calls.

Also pinned: the prompt-name collision that broke panel revision (and the
lint rule that makes the collision class impossible), spreadsheet formula
neutralization, the gitignore gap, and crash-report redaction.
"""
from __future__ import annotations

import sys
import types
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from deckscope.config import ProviderConfig


class _SpyHosted:
    """Looks like a hosted provider; records every call it receives."""

    name = "anthropic"

    def __init__(self):
        self.config = ProviderConfig(name="anthropic")
        self.calls = []

    def complete(self, *a, **k):
        self.calls.append(("complete", a))
        raise AssertionError("a hosted provider must never be reached "
                             "under NDA")

    def complete_json(self, *a, **k):
        self.calls.append(("complete_json", a))
        raise AssertionError("a hosted provider must never be reached "
                             "under NDA")


# ------------------------------------------- P0: the library path fails closed

_EXTRACTION = {"company": {"name": "Sealed Deck Co"},
               "market": {"category": "confidential widgets"},
               "claims": [{"id": "C1", "claim": "revenue is secret",
                           "load_bearing": "high"}]}


def test_engine_refuses_a_hosted_provider_before_any_call():
    """The audit drove run_research directly with a hosted provider and a
    frozen corpus: the guard refused web research, then the reader called
    the hosted model anyway with deck content in the prompt. The engine
    itself now refuses — zero outbound calls, not zero-minus-one."""
    from deckscope.research.engine import run_research
    from deckscope.research.web_backends import NoResearcher
    from deckscope.tiering import NDAGuard, NDAViolation

    provider = _SpyHosted()
    try:
        run_research(extraction=dict(_EXTRACTION), provider=provider,
                     researcher=NoResearcher(),
                     guard=NDAGuard(enabled=True),
                     deck_text="Sealed Deck Co confidential revenue")
        raise AssertionError("run_research must raise NDAViolation")
    except NDAViolation as e:
        assert "not local" in str(e)
        assert "Nothing was sent" in str(e)
    assert provider.calls == [], (
        "the refusal must come before the first provider call — a refusal "
        "after one call is a leak with an apology attached")


def test_engine_gate_accepts_local_providers():
    """mock and manual are local (the human is the boundary for manual);
    the gate must not break the offline paths."""
    from deckscope.tiering import is_local

    assert is_local(ProviderConfig(name="mock"))
    assert is_local(ProviderConfig(name="manual"))
    assert not is_local(ProviderConfig(name="anthropic"))


def test_engine_gate_refuses_a_provider_with_no_config():
    """No config means locality cannot be established — under NDA,
    unknown is refused, not trusted."""
    from deckscope.research.engine import run_research
    from deckscope.research.web_backends import NoResearcher
    from deckscope.tiering import NDAGuard, NDAViolation

    bare = types.SimpleNamespace(name="mystery")  # no .config at all
    try:
        run_research(extraction=dict(_EXTRACTION), provider=bare,
                     researcher=NoResearcher(),
                     guard=NDAGuard(enabled=True))
        raise AssertionError("must refuse a provider whose locality is "
                             "unknowable")
    except NDAViolation:
        pass


# --------------------------------- P0: improve validates BOTH providers

def test_improve_nda_refuses_a_hosted_extraction_provider(tmp_path,
                                                          monkeypatch):
    """Local main model + hosted extraction model passed the old check;
    the pipeline then sent the complete deck to the hosted one."""
    import deckscope.config as config_mod
    from deckscope.commands.improve import command

    deck = tmp_path / "notes.txt"
    deck.write_text("Our market is $47B", encoding="utf-8")

    real_load = config_mod.load_config

    def load_with_hosted_extractor(path=None, **kw):
        cfg = real_load(path, **kw)
        cfg.provider.name = "mock"
        cfg.extract_provider = ProviderConfig(name="anthropic")
        return cfg

    monkeypatch.setattr(config_mod, "load_config",
                        load_with_hosted_extractor)
    args = types.SimpleNamespace(demo=False, nda=True, lens=None,
                                 deck=str(deck), from_run=None,
                                 provider=None, config=None,
                                 out=str(tmp_path), pptx=False)
    assert command(args) == 4


# --------------------------------------------- P1: batch gets an NDA gate

def test_batch_nda_refuses_hosted_before_reading_any_deck(tmp_path,
                                                          monkeypatch):
    import deckscope.providers.registry as reg
    from deckscope.commands.batch import command

    (tmp_path / "inbound.txt").write_text("Sealed Deck Co", encoding="utf-8")
    built = []
    monkeypatch.setattr(reg, "get_provider",
                        lambda cfg: built.append(cfg))
    args = types.SimpleNamespace(folder=str(tmp_path), provider="anthropic",
                                 config=None, out=str(tmp_path / "out"),
                                 lens=None, format=None, nda=True)
    assert command(args) == 4
    assert built == [], "refusal comes before any provider is constructed"


def test_batch_nda_is_a_declared_flag():
    from deckscope.cli import build_parser

    args = build_parser().parse_args(["batch", "somedir", "--nda"])
    assert args.nda is True


# ------------------------------------- P1: the prompt collision, both halves

def test_panel_and_deck_revision_prompts_are_distinct():
    """The deck-revision prompts were appended under the panel's names;
    the second assignment won at import and panel revision died on
    KeyError('brief'). Now they are distinct names with distinct
    placeholders — and the panel's template must never need 'brief'."""
    from deckscope.prompts.templates import (DECK_REVISE_SYSTEM,
                                             DECK_REVISE_USER,
                                             REVISE_SYSTEM, REVISE_USER)

    assert "panel member" in REVISE_SYSTEM
    assert "Deck Reviser" in DECK_REVISE_SYSTEM
    assert "{brief}" in DECK_REVISE_USER
    assert "{brief}" not in REVISE_USER
    assert "{lens_block}" in REVISE_SYSTEM, (
        "the panel's system prompt keeps its own placeholder contract")


def test_linter_catches_duplicate_toplevel_assignment(tmp_path):
    """The class-level defense: a module constant assigned twice is a lint
    error, so this collision cannot be reintroduced anywhere."""
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent
                           / "scripts"))
    import lint as lint_mod

    bad = tmp_path / "bad_module.py"
    bad.write_text('X = "first"\nY = 2\nX = "second wins"\n',
                   encoding="utf-8")
    problems = lint_mod.check(bad)
    assert any("assigned twice at module level" in msg
               for (_p, _l, msg) in problems)

    ok = tmp_path / "ok_module.py"
    ok.write_text("try:\n    import json as X\nexcept ImportError:\n"
                  "    X = None\nY = 1\n", encoding="utf-8")
    assert not [m for (_p, _l, m) in lint_mod.check(ok)
                if "assigned twice" in m], (
        "conditional fallback rebinding is not a duplicate")


# ------------------------------------------ P2: spreadsheet cell neutralization

def test_formula_prefixed_cells_are_neutralized():
    from deckscope.commands.batch import neutralize_cell

    assert neutralize_cell("=HYPERLINK(evil)") == "'=HYPERLINK(evil)"
    assert neutralize_cell("+1+1") == "'+1+1"
    assert neutralize_cell("-2+3") == "'-2+3"
    assert neutralize_cell("@SUM(A1)") == "'@SUM(A1)"
    assert neutralize_cell("\t=cmd") == "'\t=cmd"
    assert neutralize_cell("Acme Flow") == "Acme Flow"
    assert neutralize_cell("— no verdict") == "— no verdict", (
        "an em dash is not a formula trigger")
    assert neutralize_cell(7) == 7


def test_write_table_applies_neutralization(tmp_path):
    from deckscope.commands.batch import write_table

    row = {"company": "=2+2", "deck": "d.pdf", "verdict": "LEAN NO",
           "confidence": "low", "contested": 1, "supported": 0,
           "unverifiable": 0, "questions": 0, "sources_cited": 0,
           "out_dir": "", "error": "@shell"}
    table = write_table([row], tmp_path)
    if table.suffix == ".xlsx":
        from openpyxl import load_workbook

        ws = load_workbook(table).active
        values = [c.value for c in ws[2]]
    else:
        import csv

        with open(table, newline="", encoding="utf-8") as fh:
            values = list(csv.reader(fh))[1]
    assert "'=2+2" in values and "'@shell" in values
    assert "=2+2" not in values


# --------------------------------------------------- P2/P3: hygiene findings

def test_gitignore_covers_the_actual_default_output_dirs():
    text = (Path(__file__).resolve().parent.parent
            / ".gitignore").read_text(encoding="utf-8")
    for entry in ("deckscope_out/", "deckscope_output/",
                  "deckscope_demo_output/", "deck_analysis/"):
        assert entry in text, (
            f"{entry} is a default output location and can hold "
            "confidential analysis — it must never be one `git add -A` "
            "away from a public repo")


def test_crash_report_redacts_home_and_warns_before_sharing(tmp_path,
                                                            monkeypatch):
    from deckscope import cli, settings

    monkeypatch.setattr(settings, "app_dir", lambda: tmp_path)
    monkeypatch.setattr(
        cli.sys, "argv",
        ["deckscope", "run", str(Path.home() / "SecretCo_deck.pdf")])
    try:
        raise RuntimeError(f"boom in {Path.home() / 'work'}")
    except RuntimeError as e:
        path = cli._crash_report(e)
    assert path is not None
    body = Path(path).read_text(encoding="utf-8")
    assert str(Path.home()) not in body, "the username travels in no path"
    assert "~" in body
    assert "review before sharing" in body
